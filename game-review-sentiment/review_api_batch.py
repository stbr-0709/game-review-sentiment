# -*- coding: utf-8 -*-
import openai
import sys
import io
import time
import json
import os

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 配置 API
API_KEY = "sk-8bf021c3ab5d42fb94e6f281644ac161"  # 👈 替换成你的真实 Key

# 如果需要代理，取消注释
# os.environ["HTTP_PROXY"] = "http://127.0.0.1:59527"
# os.environ["HTTPS_PROXY"] = "http://127.0.0.1:59527"

client = openai.OpenAI(
    api_key=API_KEY,
    base_url="https://api.deepseek.com"
)

def analyze_single_review_api(review_text):
    """调用 API 分析单条评论"""
    
    prompt = f"""分析下面的游戏用户评论，输出JSON格式。

评论："{review_text}"

只输出JSON，不要有其他内容：
{{"sentiment": "正面/负面/中性", "reason": "判断原因（一句话）", "issues": ["问题1", "问题2"]}}"""

    try:
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "你是游戏用户反馈分析专家，只输出JSON格式。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3
        )
        
        result = response.choices[0].message.content
        return result
    except Exception as e:
        return f'{{"sentiment": "分析失败", "reason": "{str(e)}", "issues": []}}'

def batch_analyze_api(reviews):
    """批量分析评论"""
    results = []
    total = len(reviews)
    
    for i, review in enumerate(reviews, 1):
        if len(review.strip()) < 3:
            continue
        
        print(f"  🤖 分析中 {i}/{total}...")
        result_json = analyze_single_review_api(review)
        
        try:
            start = result_json.find("{")
            end = result_json.rfind("}") + 1
            if start != -1 and end != 0:
                json_str = result_json[start:end]
                data = json.loads(json_str)
                sentiment = data.get("sentiment", "中性")
                reason = data.get("reason", "")
                issues = data.get("issues", [])
            else:
                sentiment = "中性"
                reason = "解析失败"
                issues = []
        except:
            sentiment = "中性"
            reason = "解析失败"
            issues = []
        
        results.append({
            "评论": review,
            "情绪": sentiment,
            "原因": reason,
            "问题": ", ".join(issues) if issues else "无"
        })
        
        time.sleep(0.3)
    
    return results

def main():
    print("=" * 55)
    print("游戏评论批量分析工具 v3.0（AI智能版）")
    print("支持：直接粘贴 / Excel导入 / 示例数据")
    print("=" * 55)
    
    while True:
        print("\n请选择模式：")
        print("1. 直接粘贴多条评论（每行一条）")
        print("2. 从 Excel/CSV 文件导入")
        print("3. 使用示例数据测试")
        print("4. 退出")
        
        choice = input("\n请输入选项（1-4）：")
        
        if choice == "1":
            print("\n📝 请输入评论（每行一条，输入空行结束）：")
            print("（可以一次性粘贴多行，然后按两次回车）")
            lines = []
            while True:
                line = input()
                if line == "":
                    break
                lines.append(line)
            
            if not lines:
                print("没有输入任何评论")
                continue
            
            print(f"\n🤖 开始分析 {len(lines)} 条评论...")
            results = batch_analyze_api(lines)
            
        elif choice == "2":
            print("\n📁 支持格式：.csv 或 .xlsx")
            print("文件的第一列应该是评论内容")
            file_path = input("请输入文件完整路径（如：C:\\Users\\Nancy\\Desktop\\comments.xlsx）：")
            
            if not os.path.exists(file_path):
                print("文件不存在，请检查路径")
                continue
            
            try:
                if file_path.endswith(".csv"):
                    import csv
                    with open(file_path, 'r', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        reviews = [row[0] for row in reader if row]
                else:
                    # 需要先安装 openpyxl
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, data_only=True)
                    ws = wb.active
                    reviews = []
                    for row in ws.iter_rows(values_only=True):
                        if row[0]:
                            reviews.append(str(row[0]))
                
                # 跳过表头（如果第一行是标题）
                if reviews and ("评论" in reviews[0] or "内容" in reviews[0] or "feedback" in reviews[0].lower()):
                    reviews = reviews[1:]
                
                print(f"✅ 成功读取 {len(reviews)} 条评论")
                print(f"\n🤖 开始分析...")
                results = batch_analyze_api(reviews)
                
            except Exception as e:
                print(f"读取文件失败：{e}")
                print("提示：Excel文件需要先安装 openpyxl：pip install openpyxl")
                continue
            
        elif choice == "3":
            sample_reviews = [
                "这游戏真好玩，画面精美，剧情也很吸引人！",
                "太卡了，而且氪金太严重，玩不下去",
                "希望能优化一下匹配机制，经常遇到挂机队友",
                "良心游戏，不花钱也能玩得很开心，推荐！",
                "垃圾游戏，全是bug，已经卸载了",
                "虽然有点卡，但总体来说挺好玩的",
                "策划脑子有坑吗？这平衡性怎么做的",
                "活动福利挺好的，希望多出一些",
                "抽卡概率太低了，充了500块什么都没抽到",
                "服务器经常掉线，体验很差，但客服态度不错"
            ]
            print(f"\n📋 示例数据（{len(sample_reviews)}条）：")
            for i, r in enumerate(sample_reviews, 1):
                display = r[:40] + "..." if len(r) > 40 else r
                print(f"{i}. {display}")
            
            print(f"\n🤖 开始分析...")
            results = batch_analyze_api(sample_reviews)
            
        elif choice == "4":
            print("再见！")
            break
        else:
            print("无效选项")
            continue
        
        # 统计结果
        print("\n" + "=" * 55)
        print("📊 分析结果")
        print("=" * 55)
        
        sentiment_count = {"正面": 0, "负面": 0, "中性": 0}
        for r in results:
            sentiment_count[r["情绪"]] = sentiment_count.get(r["情绪"], 0) + 1
        
        total = len(results)
        print(f"\n共分析 {total} 条评论")
        print(f"😊 正面：{sentiment_count.get('正面', 0)} 条 ({sentiment_count.get('正面', 0)/total*100:.1f}%)")
        print(f"😞 负面：{sentiment_count.get('负面', 0)} 条 ({sentiment_count.get('负面', 0)/total*100:.1f}%)")
        print(f"😐 中性：{sentiment_count.get('中性', 0)} 条 ({sentiment_count.get('中性', 0)/total*100:.1f}%)")
        
        # 显示前5条详细结果
        print("\n📝 详细结果（前5条）：")
        print("-" * 55)
        for r in results[:5]:
            emoji = "😊" if r["情绪"] == "正面" else ("😞" if r["情绪"] == "负面" else "😐")
            print(f"{emoji} [{r['情绪']}] {r['评论'][:50]}...")
            print(f"   原因：{r['原因']}")
        
        # 保存结果
        import csv
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_file = f"AI分析结果_{timestamp}.csv"
        
        with open(output_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=["评论", "情绪", "原因", "问题"])
            writer.writeheader()
            writer.writerows(results)
        
        print(f"\n💾 完整结果已保存：{output_file}")
        print("\n✅ 分析完成！")
        
        cont = input("\n是否继续？（y/n）：")
        if cont.lower() != "y":
            print("再见！")
            break

if __name__ == "__main__":
    main()